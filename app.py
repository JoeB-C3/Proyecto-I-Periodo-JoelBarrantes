from flask import Flask, render_template, request, redirect, url_for, flash
import os
import openpyxl

# Inicialización y configuración de la aplicación Flask
app = Flask(__name__)
app.secret_key = 'proyecto_etapa2_secreto'

# Credenciales de acceso globales para el control de sesiones
usurio_correcto = 'admin'
contrasena_correcta = '1234'

# Definición del archivo de almacenamiento (Base de datos Excel)
EXCEL_FILE = 'contactos.xlsx'

#FUNCIONES DE ASISTENCIA Y MANEJO DE ARCHIVOS

def inicializar_excel():
    """Verifica la existencia del archivo Excel y crea la estructura inicial si no existe"""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contactos"
        ws.append(["Nombre", "Apellido", "Teléfono", "Correo", "Dirección", "Categoría", "Favorito"])
        wb.save(EXCEL_FILE)
        wb.close()

def obtener_todos_los_contactos():
    """Recupera todos los registros de contactos desde el archivo de datos"""
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active
    lista_contactos = []
    
    for num_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if fila and fila[0] is not None:
            # Control de seguridad
            fav_celda = str(fila[6]).strip().lower() if len(fila) > 6 and fila[6] else 'normal'
            es_favorito = 'favorito' if fav_celda in ['favorito', 'sí', 'si'] else 'normal'

            contacto = {
                'id': num_fila,
                'nombre': fila[0],
                'apellido': fila[1],
                'telefono': fila[2],
                'correo': fila[3],
                'direccion': fila[4],
                'categoria': fila[5] if len(fila) > 5 else 'Otro',
                'favorito': es_favorito
            }
            lista_contactos.append(contacto)
            
    wb.close()
    return lista_contactos

#RUTAS Y CONTROLADORES DE LA INTERFAZ WEB

@app.route('/', methods=['GET', 'POST'])
def inicio():
    """Ruta para la gestión y validación del inicio de sesión (Login)"""
    if request.method == 'POST':
        usuario = request.form.get('usuario')
        contrasena = request.form.get('contrasena')

        if usuario == usurio_correcto and contrasena == contrasena_correcta:
            return redirect(url_for('principal'))
        else:
            flash('Usuario o contraseña incorrectos. Intente de nuevo', 'error')
            return redirect(url_for('inicio'))
        
    return render_template('index.html')


@app.route('/principal')
def principal():
    """Ruta para la pantalla principal que lista la agenda de contactos"""
    todos_los_contactos = obtener_todos_los_contactos()
    return render_template('principal.html', contactos=todos_los_contactos)


@app.route('/agregar_contacto', methods=['GET', 'POST'])
def agregar_contacto():
    """Ruta para registrar y dar de alta un nuevo contacto en el sistema"""
    if request.method == 'POST':
        nombre = request.form.get('txtNombre')
        apellido = request.form.get('txtApellido')
        telefono = request.form.get('txtTelefono')
        correo = request.form.get('txtCorreo')
        direccion = request.form.get('txtDireccion')
        categoria = request.form.get('txtCategoria')
        
        # Guardado estandarizado del formulario de creación
        chk = request.form.get('chkFavorito')
        favorito = 'favorito' if chk in ['favorito', 'Sí', 'on'] else 'normal'

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([nombre, apellido, telefono, correo, direccion, categoria, favorito])
        wb.save(EXCEL_FILE)
        wb.close()

        flash('¡Contacto agregado exitosamente!', 'success')
        return redirect(url_for('principal'))

    return render_template('agregar_contacto.html')


@app.route('/contacto/<int:row_id>', methods=['GET', 'POST'])
def detalle_contacto(row_id):
    """Ruta para visualizar el detalle individual o editar un contacto específico"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # SI EL USUARIO ENVÍA EL FORMULARIO (EDICIÓN)
    if request.method == 'POST':
        ws.cell(row=row_id, column=1).value = request.form.get('txtNombre')
        ws.cell(row=row_id, column=2).value = request.form.get('txtApellido')
        ws.cell(row=row_id, column=3).value = request.form.get('txtTelefono')
        ws.cell(row=row_id, column=4).value = request.form.get('txtCorreo')
        ws.cell(row=row_id, column=5).value = request.form.get('txtDireccion')
        ws.cell(row=row_id, column=6).value = request.form.get('txtCategoria')
        
        # Captura directa del radio seleccionado ('favorito' o 'normal')
        fav_form = request.form.get('chkFavorito', 'normal')
        ws.cell(row=row_id, column=7).value = 'favorito' if fav_form in ['favorito', 'Sí'] else 'normal'
        
        wb.save(EXCEL_FILE)
        wb.close()
        flash('Contacto actualizado con éxito', 'success')
        return redirect(url_for('principal'))

    # SI EL USUARIO SOLO ENTRA A VER LA INFORMACIÓN (GET)
    fav_celda = str(ws.cell(row=row_id, column=7).value).strip().lower()
    es_favorito = 'favorito' if fav_celda in ['favorito', 'sí', 'si'] else 'normal'

    contacto = {
        'id': row_id,
        'nombre': ws.cell(row=row_id, column=1).value,
        'apellido': ws.cell(row=row_id, column=2).value,
        'telefono': ws.cell(row=row_id, column=3).value,
        'correo': ws.cell(row=row_id, column=4).value,
        'direccion': ws.cell(row=row_id, column=5).value,
        'categoria': ws.cell(row=row_id, column=6).value,
        'favorito': es_favorito
    }
    wb.close()
    return render_template('detalle_contacto.html', contacto=contacto)


@app.route('/contacto/eliminar/<int:row_id>', methods=['POST'])
def eliminar_contacto(row_id):
    """Ruta para la eliminación permanente de un registro en el Excel"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    ws.delete_rows(row_id, 1)
    
    wb.save(EXCEL_FILE)
    wb.close()
    flash('Contacto eliminado satisfactoriamente', 'success')
    return redirect(url_for('principal'))


@app.route('/favoritos')
def favoritos():
    """Ruta encargada de filtrar y listar solo los contactos preferidos"""
    todos = obtener_todos_los_contactos()
    lista_favoritos = [c for c in todos if c['favorito'] == 'favorito']
    return render_template('favoritos.html', contactos=lista_favoritos)


@app.route('/reporte_general')
def reporte_general():
    """Ruta orientada a calcular las estadísticas e indicadores generales del sistema"""
    # 1. Obtenemos la lista con todos los contactos desde el Excel
    todos = obtener_todos_los_contactos()
    
    # 2. Contamos cuántos elementos hay en total
    total_contactos = len(todos)
    
    # 3. Filtramos y contamos cuántos están marcados como 'favorito'
    total_favoritos = len([c for c in todos if c['favorito'] == 'favorito'])
    
   #se renderiza la pagina junto con las variables necesarias para el código
    return render_template('reporte_general.html', total_contactos=total_contactos, total_favoritos=total_favoritos)

# --- INICIALIZADOR DEL SERVIDOR ---
if __name__ == '__main__':
    app.run(debug=True)